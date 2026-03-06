"""
track_meeting_declines.py
--------------------------
Uses pywin32 to open Outlook, find meetings matching a keyword,
track all attendee responses (including decline reasons from reply emails),
and export results to a formatted Excel workbook with:
  - "Summary"   sheet : pivot-style attendance grid (person x date)
  - "Responses" sheet : per-occurrence detail (name, email, status, decline reason)

Requirements:
    pip install pywin32 openpyxl

Usage:
    python track_meeting_declines.py
    python track_meeting_declines.py --subject "Weekly Standup"
    python track_meeting_declines.py --subject "Weekly Standup" --days 90
    python track_meeting_declines.py --subject "Weekly Standup" --days 90 --out report.xlsx
    python track_meeting_declines.py --subject "Weekly Standup" --no-reasons
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import win32com.client
except ImportError:
    print("ERROR: pywin32 is not installed.  Run:  pip install pywin32")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed.  Run:  pip install openpyxl")
    sys.exit(1)


# ── Outlook constants ────────────────────────────────────────────────────────
OL_FOLDER_CALENDAR = 9
OL_FOLDER_INBOX    = 6

OL_RESPONSE = {
    3: "Accepted",
    4: "Declined",
    2: "Tentative",
    5: "No Response",
}

STATUS_ICON = {
    "Accepted":    "✔",
    "Declined":    "✘",
    "Tentative":   "~",
    "No Response": "?",
    "—":           "—",
}

STATUS_FILL = {
    "Accepted":    "C6EFCE",
    "Declined":    "FFC7CE",
    "Tentative":   "FFEB9C",
    "No Response": "D9D9D9",
}


# ── Outlook helpers ──────────────────────────────────────────────────────────

def connect_outlook():
    try:
        app = win32com.client.Dispatch("Outlook.Application")
        print("✔  Connected to Outlook.")
        return app
    except Exception as e:
        print(f"ERROR connecting to Outlook: {e}")
        sys.exit(1)


def get_folder(outlook, folder_id: int):
    return outlook.GetNamespace("MAPI").GetDefaultFolder(folder_id)


# ── Decline-reason extraction ────────────────────────────────────────────────

def _strip_quoted(text: str) -> str:
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if s.startswith(">"):
            continue
        if re.match(r"^-{5,}", s) or re.match(r"^_{5,}", s):
            break
        lines.append(line)
    return "\n".join(lines).strip()


def _clean_reason(text: str) -> str:
    text = _strip_quoted(text)
    text = re.sub(r"\s+", " ", text).strip()
    auto = [r"^accepted\b", r"^declined\b", r"^tentative\b", r"microsoft outlook"]
    for pat in auto:
        if re.match(pat, text, re.IGNORECASE):
            return ""
    return text[:200] if text else ""


def find_decline_reason(inbox, attendee_email: str, meeting_subject: str,
                        meeting_start: datetime) -> str:
    window_start = meeting_start - timedelta(days=7)
    window_end   = meeting_start + timedelta(days=7)
    filter_str = (
        f"[ReceivedTime] >= '{window_start.strftime('%m/%d/%Y')}' AND "
        f"[ReceivedTime] <= '{window_end.strftime('%m/%d/%Y')}'"
    )
    try:
        items = inbox.Items.Restrict(filter_str)
    except Exception:
        return ""

    subj_lower   = meeting_subject.lower()
    sender_lower = attendee_email.lower()

    for item in items:
        try:
            msg_class = (getattr(item, "MessageClass", "") or "").lower()
            if "resp.neg" not in msg_class:
                continue
            sender = (getattr(item, "SenderEmailAddress", "") or "").lower()
            subj   = (getattr(item, "Subject", "") or "").lower()
            if sender_lower not in sender or subj_lower not in subj:
                continue
            body = getattr(item, "Body", "") or ""
            reason = _clean_reason(body)
            if reason:
                return reason
        except Exception:
            continue
    return ""


# ── Calendar search ──────────────────────────────────────────────────────────

def find_meetings(calendar, subject_kw: str, since: datetime):
    items = calendar.Items
    items.IncludeRecurrences = True
    items.Sort("[Start]")
    results = []
    for item in items:
        try:
            if subject_kw.lower() not in (item.Subject or "").lower():
                continue
            s = item.Start
            py_start = datetime(s.year, s.month, s.day, s.hour, s.minute)
            if py_start >= since:
                results.append((py_start, item))
        except Exception:
            continue
    return results


# ── Data collection ──────────────────────────────────────────────────────────

def collect_data(meetings, inbox, fetch_reasons: bool):
    rows   = []
    dates  = set()
    people = {}   # email -> name

    for idx, (py_start, item) in enumerate(meetings, 1):
        ds   = py_start.strftime("%Y-%m-%d")
        subj = item.Subject or ""
        print(f"  [{idx}/{len(meetings)}] {ds}  {subj}")

        for i in range(1, item.Recipients.Count + 1):
            try:
                r      = item.Recipients.Item(i)
                name   = r.Name or "Unknown"
                email  = (r.Address or "").lower()
                status = OL_RESPONSE.get(r.MeetingResponseStatus, "No Response")
                reason = ""
                if status == "Declined" and fetch_reasons and inbox:
                    reason = find_decline_reason(inbox, email, subj, py_start)
                rows.append({
                    "date": ds, "subject": subj,
                    "name": name, "email": email,
                    "status": status, "reason": reason,
                })
                dates.add(py_start.date())
                if email:
                    people[email] = name
            except Exception:
                continue

    return rows, sorted(dates), people


# ── Shared style helpers ─────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center")


def hdr(cell, value):
    cell.value     = value
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def col_w(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Responses sheet ──────────────────────────────────────────────────────────

def build_responses_sheet(wb, rows):
    ws = wb.create_sheet("Responses")
    ws.freeze_panes = "A2"

    headers = ["Date", "Meeting Subject", "Attendee Name",
               "Email", "Status", "Decline Reason"]
    widths  = [13,      40,               25,
               34,      13,      60]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        hdr(ws.cell(1, col), h)
        col_w(ws, col, w)
    ws.row_dimensions[1].height = 20

    for ri, row in enumerate(rows, 2):
        alt  = ri % 2 == 0
        vals = [row["date"], row["subject"], row["name"],
                row["email"], row["status"], row["reason"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val)
            c.font   = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(ci == 6))
            if ci == 5:
                c.fill = PatternFill("solid",
                                     fgColor=STATUS_FILL.get(val, "FFFFFF"))
                c.alignment = CENTER
            elif alt:
                c.fill = ALT_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Summary sheet ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, rows, sorted_dates, people):
    ws = wb.create_sheet("Summary", 0)
    ws.freeze_panes = "B2"

    lookup = {}
    for row in rows:
        lookup[(row["email"], row["date"])] = (row["status"], row["reason"])

    date_strs  = [d.strftime("%Y-%m-%d") for d in sorted_dates]
    n_dates    = len(date_strs)
    all_emails = sorted(people.keys(), key=lambda e: people[e].lower())
    n_people   = len(all_emails)

    stat_cols = ["Accepted", "Declined", "Tentative", "No Response", "Attendance %"]
    stat_start = n_dates + 2   # first summary-stat column index (1-based)

    # ── Row 1: headers ────────────────────────────────────────────────────────
    hdr(ws.cell(1, 1), "Attendee")
    col_w(ws, 1, 28)

    for ci, ds in enumerate(date_strs, 2):
        hdr(ws.cell(1, ci), ds)
        col_w(ws, ci, 13)

    for off, label in enumerate(stat_cols):
        hdr(ws.cell(1, stat_start + off), label)
        col_w(ws, stat_start + off, 14)

    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, email in enumerate(all_emails, 2):
        c = ws.cell(ri, 1, people[email])
        c.font   = BOLD_FONT
        c.border = BORDER
        c.alignment = Alignment(vertical="center")

        for ci, ds in enumerate(date_strs, 2):
            status, reason = lookup.get((email, ds), ("—", ""))
            icon = STATUS_ICON.get(status, "—")
            cell = ws.cell(ri, ci, icon)
            cell.font      = BODY_FONT
            cell.border    = BORDER
            cell.alignment = CENTER
            fgColor = STATUS_FILL.get(status, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fgColor)
            if reason:
                cell.comment = Comment(
                    f"Decline reason:\n{reason}", "Outlook Tracker"
                )

        # COUNTIF formulas over the icon columns
        data_rng = (f"{get_column_letter(2)}{ri}:"
                    f"{get_column_letter(n_dates + 1)}{ri}")

        for off, key in enumerate(["Accepted", "Declined", "Tentative", "No Response"]):
            c = ws.cell(ri, stat_start + off,
                        f'=COUNTIF({data_rng},"{STATUS_ICON[key]}")')
            c.font = BODY_FONT; c.border = BORDER; c.alignment = CENTER

        # Attendance %  (Accepted / total occurrences)
        acc_ref = f"{get_column_letter(stat_start)}{ri}"
        pct = ws.cell(ri, stat_start + 4,
                      f"=IF({n_dates}=0,0,{acc_ref}/{n_dates})")
        pct.number_format = "0.0%"
        pct.font = BODY_FONT; pct.border = BORDER; pct.alignment = CENTER

    # ── Totals row ────────────────────────────────────────────────────────────
    tr = n_people + 2
    c = ws.cell(tr, 1, "TOTAL")
    c.font = BOLD_FONT; c.border = BORDER
    TOTALS_FILL = PatternFill("solid", fgColor="D9E1F2")

    for ci in range(2, n_dates + 2):
        cl = get_column_letter(ci)
        c = ws.cell(tr, ci,
                    f'=COUNTIF({cl}2:{cl}{tr-1},"{STATUS_ICON["Accepted"]}")')
        c.font = BOLD_FONT; c.border = BORDER
        c.alignment = CENTER; c.fill = TOTALS_FILL

    for off in range(4):
        cl = get_column_letter(stat_start + off)
        c = ws.cell(tr, stat_start + off, f"=SUM({cl}2:{cl}{tr-1})")
        c.font = BOLD_FONT; c.border = BORDER
        c.alignment = CENTER; c.fill = TOTALS_FILL

    acc_cl = get_column_letter(stat_start)
    pct = ws.cell(tr, stat_start + 4,
                  f"=IF({n_dates}*{n_people}=0,0,"
                  f"SUM({acc_cl}2:{acc_cl}{tr-1})/({n_dates}*{n_people}))")
    pct.number_format = "0.0%"
    pct.font = BOLD_FONT; pct.border = BORDER
    pct.alignment = CENTER; pct.fill = TOTALS_FILL

    # ── Legend ────────────────────────────────────────────────────────────────
    lr = tr + 2
    ws.cell(lr, 1, "Legend:").font = Font(name="Arial", bold=True, size=9)
    for i, (key, label) in enumerate(
        [("Accepted", "Accepted"), ("Declined", "Declined"),
         ("Tentative", "Tentative"), ("No Response", "No Response")]
    ):
        ic = ws.cell(lr, 2 + i * 2, STATUS_ICON[key])
        ic.fill = PatternFill("solid", fgColor=STATUS_FILL[key])
        ic.font = Font(name="Arial", size=9)
        ic.alignment = CENTER; ic.border = BORDER
        lc = ws.cell(lr, 3 + i * 2, label)
        lc.font = Font(name="Arial", size=9)

    # Note about hover-over comments for decline reasons
    note_row = lr + 1
    ws.cell(note_row, 1,
            "💡 Hover over a ✘ cell to see the attendee's decline reason (if captured)."
            ).font = Font(name="Arial", italic=True, size=9, color="595959")
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=min(8, n_dates + 2)
    )


# ── Export ───────────────────────────────────────────────────────────────────

def export_excel(rows, sorted_dates, people, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("\n📊 Building Summary sheet …")
    build_summary_sheet(wb, rows, sorted_dates, people)

    print("📋 Building Responses sheet …")
    build_responses_sheet(wb, rows)

    wb.save(out_path)
    print(f"\n✅  Saved: {out_path.resolve()}")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Track Outlook meeting responses and export to Excel."
    )
    parser.add_argument("--subject", default="",
                        help="Keyword to match in meeting subject (case-insensitive).")
    parser.add_argument("--days", type=int, default=90,
                        help="How many past days to search (default: 90).")
    parser.add_argument("--out", default="meeting_responses.xlsx",
                        help="Output Excel file (default: meeting_responses.xlsx).")
    parser.add_argument("--no-reasons", action="store_true",
                        help="Skip inbox search for decline reasons (faster).")
    args = parser.parse_args()

    if not args.subject:
        args.subject = input("Meeting subject keyword: ").strip()
        if not args.subject:
            print("No subject entered. Exiting.")
            sys.exit(0)

    since = datetime.now() - timedelta(days=args.days)

    outlook  = connect_outlook()
    calendar = get_folder(outlook, OL_FOLDER_CALENDAR)
    inbox    = None if args.no_reasons else get_folder(outlook, OL_FOLDER_INBOX)

    print(f"\n🔍 Searching for '{args.subject}' in the last {args.days} days …")
    meetings = find_meetings(calendar, args.subject, since)

    if not meetings:
        print("❌  No matching meetings found. Try a different keyword or --days value.")
        sys.exit(1)

    print(f"   Found {len(meetings)} occurrence(s). Collecting responses …\n")
    rows, sorted_dates, people = collect_data(
        meetings, inbox, fetch_reasons=not args.no_reasons
    )

    if not rows:
        print("No attendee data found.")
        sys.exit(1)

    export_excel(rows, sorted_dates, people, Path(args.out))


if __name__ == "__main__":
    main()
