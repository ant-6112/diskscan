from utils.config_utils import get_full_config
from utils.log_utils import call_logger

from pathlib import Path
import pandas as pd
import os
import warnings
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import pymupdf as pm

warnings.filterwarnings("ignore")

config = get_full_config()["config"]

def read_extraction_report(workbook_path):
    sheet_id = int(config["extraction_sheet_id"])
    path = Path(workbook_path)

    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found at: {path}")

    df = pd.read_excel(path, sheet_name=sheet_id, dtype=str)
    if df.empty:
        return {}

    # Common pattern: 2-column key/value sheet
    if df.shape[1] >= 2:
        key_col = df.columns[0]
        val_col = df.columns[1]
        keys = df[key_col].astype(str).str.strip().tolist()
        vals = df[val_col].astype(str).str.strip().tolist()
        return dict(zip(keys, vals))

    # Fallback: single-row table
    records = df.astype(str).to_dict(orient="records")
    return records[0] if records else {}

def read_all_reports():
    currencies = []
    reports_folder = Path("Inputs/Reports")
    if not reports_folder.is_dir():
        raise FileNotFoundError(f"Reports folder not found at: {reports_folder}")

    all_reports = {}
    for name in os.listdir(reports_folder):
        if not name.lower().endswith(".xlsx"):
            continue
        workbook_path = reports_folder / name

        if not workbook_path.is_file():
            continue

        temp_df = read_extraction_report(workbook_path=workbook_path)
        if not temp_df:
            continue

        interest_amount = (temp_df.get("Interest Amount") or "").strip()
        if len(interest_amount) < 3:
            continue

        report_curr = interest_amount[0:3]
        temp_df["Currency"] = report_curr

        if report_curr not in currencies:
            currencies.append(report_curr)
            all_reports[report_curr] = []
        all_reports[report_curr].append(temp_df)

    return all_reports


def _norm_header(value):
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_header_row(ws, required_header=None, max_scan_rows=50):
    required_norm = _norm_header(required_header) if required_header else None
    max_row = min(max_scan_rows, ws.max_row or max_scan_rows)
    max_col = max(1, ws.max_column or 1)

    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            row_vals.append(v)
            if required_norm and _norm_header(v) == required_norm:
                return r
        if not required_norm:
            non_empty = sum(1 for v in row_vals if str(v).strip() != "")
            if non_empty >= 2:
                return r
    return 1


def _header_map(ws, header_row):
    max_col = max(1, ws.max_column or 1)
    mapping = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = str(v).strip()
        if key:
            mapping[key] = c
    return mapping


def _last_data_row(ws, header_row, header_cols):
    if not header_cols:
        return header_row

    last = header_row
    # Scan upward to avoid formatted-but-empty tails.
    for r in range(ws.max_row or header_row, header_row, -1):
        for c in header_cols:
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                return r
    return last


def _copy_sheet_contents(template_ws, target_ws):
    # Copy cell values + styles
    for row in template_ws.iter_rows():
        for cell in row:
            tgt = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                tgt._style = cell._style
            tgt.number_format = cell.number_format
            tgt.protection = cell.protection
            tgt.alignment = cell.alignment
            tgt.font = cell.font
            tgt.fill = cell.fill
            tgt.border = cell.border

    # Dimensions
    for col_letter, dim in template_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
        target_ws.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx, dim in template_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height
        target_ws.row_dimensions[row_idx].hidden = dim.hidden

    # Sheet view settings
    target_ws.freeze_panes = template_ws.freeze_panes
    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
    target_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight

    # Merged cells
    for merged in list(template_ws.merged_cells.ranges):
        target_ws.merge_cells(str(merged))

    # Auto filter
    if template_ws.auto_filter and template_ws.auto_filter.ref:
        target_ws.auto_filter.ref = template_ws.auto_filter.ref


def update_output_workbook_with_reports(
    new_reports_data,
    output_workbook_path,
    template_workbook_path=None,
):
    out_path = Path(output_workbook_path)
    template_path = Path(
        template_workbook_path
        if template_workbook_path is not None
        else config["template_name"]
    )

    if not template_path.is_file():
        raise FileNotFoundError(f"Template workbook not found at: {template_path}")

    template_wb = load_workbook(template_path)
    template_ws = template_wb.worksheets[0]

    if out_path.is_file():
        out_wb = load_workbook(out_path)
    else:
        # Start from an empty workbook; we will add currency sheets as needed.
        out_wb = load_workbook(template_path)
        # If we created from template, keep its existing sheets; caller may want them.
        # We'll still create currency sheets (or reuse if they already exist).

    required_header = config.get("refer_col")

    for currency, reports in (new_reports_data or {}).items():
        if not reports:
            continue

        sheet_name = str(currency).strip()
        if not sheet_name:
            continue

        if sheet_name in out_wb.sheetnames:
            ws = out_wb[sheet_name]
        else:
            ws = out_wb.create_sheet(title=sheet_name)
            _copy_sheet_contents(template_ws, ws)

        header_row = _find_header_row(ws, required_header=required_header)
        headers = _header_map(ws, header_row)
        if not headers:
            # No discernible header; skip to avoid corrupting the sheet.
            continue

        headers_norm = {_norm_header(k): k for k in headers.keys()}

        # Determine append start row
        header_cols = list(headers.values())
        last_row = _last_data_row(ws, header_row, header_cols)
        next_row = max(last_row + 1, header_row + 1)

        # Choose a style donor row (first data row if present, otherwise the row right below header)
        style_row = header_row + 1
        if last_row >= header_row + 1:
            style_row = header_row + 1

        def _get_value(report, header_name):
            if header_name in report:
                return report.get(header_name)
            norm = _norm_header(header_name)
            k = headers_norm.get(norm)
            if k and k in report:
                return report.get(k)
            # Try normalized match across report keys
            report_norm = {_norm_header(rk): rk for rk in report.keys()}
            rk = report_norm.get(norm)
            return report.get(rk) if rk else None

        for report in reports:
            if not isinstance(report, dict):
                continue

            for header_name, col in headers.items():
                val = _get_value(report, header_name)
                cell = ws.cell(row=next_row, column=col, value=val)

                donor = ws.cell(row=style_row, column=col)
                if donor.has_style:
                    cell._style = donor._style
                cell.number_format = donor.number_format
                cell.protection = donor.protection
                cell.alignment = donor.alignment
                cell.font = donor.font
                cell.fill = donor.fill
                cell.border = donor.border

            next_row += 1

        # If the sheet has an Excel Table, extend it to include newly appended rows
        final_row = next_row - 1
        try:
            tables = getattr(ws, "tables", None)
            if tables:
                for tbl in tables.values():
                    if not getattr(tbl, "ref", None):
                        continue
                    start_ref, end_ref = tbl.ref.split(":")
                    start_row = ws[start_ref].row
                    if start_row != header_row:
                        continue
                    # Keep same end column, extend end row
                    end_col = ws[end_ref].column
                    tbl.ref = f"{start_ref}:{ws.cell(row=final_row, column=end_col).coordinate}"
        except Exception:
            pass

    out_wb.save(out_path)


def add_pdf_first_pages_to_sheet(pdf_folder_path, workbook_path):
    pdf_folder = Path(pdf_folder_path)
    if not pdf_folder.is_dir():
        raise FileNotFoundError(f"PDF folder not found at: {pdf_folder}")

    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found at: {path}")

    sheet_id = int(config["extraction_sheet_id"])
    wb = load_workbook(path)
    ws = wb.worksheets[sheet_id]

    current_row = ws.max_row + 5
    row_step = 40

    for name in sorted(os.listdir(pdf_folder)):
        if not name.lower().endswith(".pdf"):
            continue

        pdf_path = pdf_folder / name
        if not pdf_path.is_file():
            continue

        doc = pm.open(pdf_path)
        if doc.page_count == 0:
            doc.close()
            continue

        page = doc.load_page(0)
        pix = page.get_pixmap()

        with NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = tmp.name
            pix.save(tmp_path)

        doc.close()

        img = XLImage(tmp_path)
        anchor_cell = f"A{current_row}"
        ws.add_image(img, anchor_cell)
        current_row += row_step

    wb.save(path)


def extract_right_of_keyword_first_page(pdf_path, keyword, case_sensitive=False):
    path = Path(pdf_path)
    if not path.is_file():
        raise FileNotFoundError(f"PDF not found at: {path}")

    if not keyword:
        return []

    needle = keyword if case_sensitive else keyword.lower()

    doc = pm.open(path)
    try:
        if doc.page_count == 0:
            return []

        page = doc.load_page(0)
        text_dict = page.get_text("dict")

        matches = []
        for block in text_dict.get("blocks", []):
            for line in block.get("lines", []):
                spans = line.get("spans", [])
                line_text = "".join(span.get("text", "") for span in spans).strip()
                if not line_text:
                    continue

                haystack = line_text if case_sensitive else line_text.lower()
                start = 0
                while True:
                    idx = haystack.find(needle, start)
                    if idx == -1:
                        break
                    right = line_text[idx + len(keyword) :].strip()
                    matches.append(right)
                    start = idx + len(needle)

        return matches
    finally:
        doc.close()


def main():
    try:
        new_reports_data = read_all_reports()
    except Exception as e:
        call_logger(e)

if __name__ == "__main__":
    main()
